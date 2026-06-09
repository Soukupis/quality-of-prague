"""XLSX data loader for ČSÚ demographic statistics.

Reads the annual demographic time series workbook published by the Czech
Statistical Office (ČSÚ) and returns per-district indicator values.

The workbook structure (all sheets identical schema):
  Row 3 (index 2): column headers — col A = indicator label, col B+ = district names
  Row 4+: indicator rows, col A = Czech/English label, col B+ = values

District names in the XLSX use "Praha-{suburb}" for suburbs, while the GeoJSON
uses just "{suburb}" (e.g., "Praha-Kunratice" → "Kunratice"). This module
normalises names before returning.
"""
from importlib import resources
from functools import lru_cache

PKG_NAME = "data"
XLSX_PATH = "demograficke_udaje_mc/demograficke_udaje_mc.xlsx"

# Row labels (partial match) → output key mapping
_INDICATOR_PATTERNS = {
    "Počet obyvatel k 31": "population",
    "Hustota zalidnění": "pop_density_per_km2",
    "0 - 14": "age_0_14_pct",
    "15 - 64": "age_15_64_pct",
    "65+": "age_65plus_pct",
    "Průměrný věk obyvatel": "mean_age",
}


def _normalise_district(xlsx_name: str) -> str:
    """Convert XLSX district name to GeoJSON district name."""
    if xlsx_name.startswith("Praha-"):
        return xlsx_name[6:]   # "Praha-Kunratice" → "Kunratice"
    return xlsx_name           # "Praha 1", "Praha 7", etc. stay as-is


@lru_cache(maxsize=4)
def get_district_demographics(year: int = 2024) -> dict:
    """Return demographic indicators for all Prague districts from ČSÚ XLSX.

    Args:
        year: Calendar year to read (sheet name). Defaults to 2024.

    Returns:
        dict: {district_name: {indicator_key: value, ...}}
        Keys: population, pop_density_per_km2, age_0_14_pct, age_15_64_pct,
              age_65plus_pct, mean_age
    """
    try:
        import openpyxl
    except ImportError:
        return {}

    ref = resources.files(PKG_NAME) / XLSX_PATH
    with resources.as_file(ref) as path:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    sheet_name = str(year)
    if sheet_name not in wb.sheetnames:
        return {}

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # Row index 2 (0-based) = district name headers
    header_row = rows[2]
    # Build col_index → normalised district name (skip col 0 = indicator label, col 1 = "Hl. m. Praha")
    col_to_district = {}
    for col_idx, cell in enumerate(header_row):
        if col_idx < 2 or cell is None:
            continue
        name = _normalise_district(str(cell).strip())
        if name not in ("Hl. m. Praha", "Ostatní \nOther", "Ostatní"):
            col_to_district[col_idx] = name

    # Initialise result dict
    result = {d: {} for d in col_to_district.values()}

    # Scan rows for matching indicators — first occurrence wins (setdefault)
    for row in rows[3:]:
        label = str(row[0]).strip() if row[0] else ""
        for pattern, key in _INDICATOR_PATTERNS.items():
            if pattern in label:
                for col_idx, district in col_to_district.items():
                    val = row[col_idx] if col_idx < len(row) else None
                    if val is not None:
                        try:
                            result[district].setdefault(key, float(val))
                        except (TypeError, ValueError):
                            pass
                break  # found match, skip other patterns for this row

    return result
